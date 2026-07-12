import os
import sys
import json
import yaml
import openpyxl
import argparse
from copy import copy
from datetime import date
from jinja2 import Environment, FileSystemLoader
from openpyxl.styles import Border, Side
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.worksheet.table import Table
from pydantic import ValidationError

from schemas import Publication, PersonRecord, Affiliation
from authors import render_authors

parser = argparse.ArgumentParser(description='Generate Website, CV, COA')
parser.add_argument('--onepager', action='store_true',
                    help='One-pager vs full')
args = parser.parse_args()

jinja_env = Environment(loader=FileSystemLoader("templates"))

# LaTeX's own syntax is full of { and }, which collide with Jinja's default
# delimiters and force every template to wrap literal LaTeX in {% raw %}/{% endraw %}.
# This environment uses LaTeX-friendly delimiters instead, so CV templates can be
# plain LaTeX with \VAR{...}/\BLOCK{...} sprinkled in and no raw blocks anywhere.
latex_env = Environment(
    block_start_string=r'\BLOCK{',
    block_end_string='}',
    variable_start_string=r'\VAR{',
    variable_end_string='}',
    comment_start_string=r'\#{',
    comment_end_string='}',
    loader=FileSystemLoader("templates"),
)

def load_yaml(path):
    return yaml.load(open(path), Loader=yaml.CLoader)

# Load Publications 
def load_validated(path, model):
    entries = load_yaml(path)
    validated = []
    for i, entry in enumerate(entries):
        try:
            validated.append(model(**entry).model_dump(exclude_unset=True))
        except ValidationError as e:
            sys.exit(f"{path}: entry {i} ({entry.get('name', entry.get('title', '?'))}) failed validation:\n{e}")
    return validated

pubs = load_validated("yaml/publications.yaml", Publication)
if args.onepager:
    pubs = [p for p in pubs if p.get("onepager")]

# Load Author urls
webs = load_yaml("yaml/websites.yaml")

# Students: one record per person, each with a list of stints (postdoc/phd/other x
# current/alumni). A person with both a current and an alumni stint (e.g. someone who
# did a Masters in the lab and is now a PhD student) shows up in both derived lists below.
people = load_validated("yaml/people.yaml", PersonRecord)

def stints_for(program, status):
    return [{**{k: v for k, v in p.items() if k != "stints"}, **stint}
            for p in people for stint in p["stints"]
            if stint["program"] == program and stint["status"] == status]

postdoc_yaml = stints_for("postdoc", "current")
student_yaml = stints_for("phd", "current")
student_names = set([s["name"] for s in student_yaml])
alumni_phd = stints_for("phd", "alumni")
masters = stints_for("other", "current")
alumni = stints_for("other", "alumni")

# The CV counts every degree, so someone who did a Masters here and is now a current
# PhD student here appears in both `alumni` and `student_yaml`. The website's alumni
# table would be redundant for them (they're already shown as a current PhD card), so
# it gets the subset that excludes anyone currently active as a PhD student here.
alumni_website = [a for a in alumni if a["name"] not in student_names]

# Theses
theses_yaml = load_yaml("yaml/theses.yaml")
teach_yaml = load_yaml("yaml/teaching.yaml")
areachair_yaml = load_yaml("yaml/area_chair.yaml")

# Load template
types = {
         "Conference": ["inproceedings", "booktitle"],
         "Journal": ["article", "journal"],
         "Preprint": ["article", "journal"],
         "Workshop": ["inproceedings", "booktitle"],
        }
colors = {
          "WS1":"primary", 
          "WS2":"primary", 
          "WS3": "success", 
          "WS4": "danger", 
          "WS5": "warning", 
          "O": "info", 
          "A": "secondary"
        }

def update_pub_entry(entry):
  # Authors and links to all co-authors
  entry["authors_pretty"] = render_authors(entry["authors"], webs, student_names)
  # Alt text for the `fig` thumbnail; a hand-written `caption` always wins.
  if entry.get("fig") and not entry.get("caption"):
      entry["caption"] = f'Key figure from {entry["title"]}'

def build_pubs_jsonld(entries, list_name):
    items = []
    for i, entry in enumerate(entries, start=1):
        article = {
            "@type": "ScholarlyArticle",
            "name": entry["title"],
            "author": [{"@type": "Person", "name": a} for a in entry["authors"]],
            "datePublished": entry["year"],
            "isPartOf": {"@type": "Periodical", "name": entry["venue"]},
        }
        if entry.get("url"):
            article["url"] = entry["url"]
        if entry.get("fig"):
            article["image"] = f"https://talkingtorobots.com/figures/{entry['fig']}"
        items.append({"@type": "ListItem", "position": i, "item": article})

    return json.dumps({
        "@context": "https://schema.org",
        "@type": "ItemList",
        "name": list_name,
        "itemListElement": items,
    }, indent=2)

def generate_publications_website():
    pub_template = jinja_env.get_template("pub_template.jinja2")

    for entry in pubs:
      update_pub_entry(entry)

    ## Generate Plot ##
    data = {"WS1":0, "WS2":0, "WS3":0, "WS4":0, "WS5":0, "O":0}
    for entry in pubs:
        u = 1.0
        data[entry["field"]] += u

    recent_pubs_jsonld = build_pubs_jsonld(pubs[:5], "Most Recent Publications - CLAW @ CMU")
    all_pubs_jsonld = build_pubs_jsonld(pubs, "Publications - CLAW @ CMU")

    rendered = pub_template.render(publications=pubs, data=data, colors=colors, types=types,
                                   recent_pubs_jsonld=recent_pubs_jsonld,
                                   all_pubs_jsonld=all_pubs_jsonld)
    with open("../publications.html", 'wt') as output_file:
        output_file.write(rendered)

def number_pub(vals, pub_count=1):  
    for val in vals:
        val["idx"] = pub_count
        pub_count += 1
    return pub_count
        
def generate_CV():
    ## Generate CV ##
    if args.onepager:
        template = "CV_template_1p.jinja2"
    else:
        template = "CV_template.jinja2"
    latex_template = latex_env.get_template(template)
    latex_papers = {"Journal": [],
                    "Conference": [],
                    "Workshop": [],
                    "Preprint": [],
                    }
    for entry in pubs:
      latex_papers[entry["type"]].append(entry)

    # Hi dear reader, why is Yonatan doing this ugly thing? 
    # This categorization and numbering is a required format for submitting 
    # promotion materials
    pub_count = number_pub(latex_papers["Journal"])
    pub_count = number_pub(latex_papers["Conference"], pub_count)
    pub_count = number_pub(latex_papers["Workshop"], pub_count)
    number_pub(latex_papers["Preprint"], pub_count)

    if args.onepager:
      latex_render = latex_template.render(pub_types=["Journal", "Conference"], 
                                           publications=latex_papers,
                                           students=student_yaml,
                                           postdocs=postdoc_yaml,
                                           teaching=teach_yaml,
                                           areachair=areachair_yaml,
                                           theses=theses_yaml)
    else:
      latex_render = latex_template.render(pub_types=["Journal", "Conference", "Workshop", "Preprint"],
                                           publications=latex_papers,
                                           students=student_yaml,
                                           postdocs=postdoc_yaml,
                                           alumni_phd=alumni_phd,
                                           masters=masters,
                                           alumni=alumni,
                                           teaching=teach_yaml,
                                           areachair=areachair_yaml,
                                           theses=theses_yaml)
    with open("CV.tex", 'wt') as output_file:
        output_file.write(latex_render)
    if os.system("pdflatex -interaction=nonstopmode CV.tex") != 0 or not os.path.exists("CV.pdf"):
        sys.exit("pdflatex failed to produce CV.pdf; see CV.log. Leaving the previous CV.pdf untouched.")
    os.system("rm CV.tex CV.log CV.aux CV.out")
    os.system("mv CV.pdf ../")

def generate_group_page():
    ## Generate group page
    group_template = jinja_env.get_template("group_template.jinja2")
    for stud in student_yaml:
        stud["research"] = []
        for pub in pubs:
            if stud["name"] in pub["authors"] and pub["type"] != "workshop":
                stud["research"].append(pub)
    group_render = group_template.render(students=student_yaml,
                                         postdocs=postdoc_yaml,
                                         alumni=alumni_website,
                                         alumni_phd=alumni_phd,
                                         masters=masters)
    with open("../CLAW/index.html", 'wt') as output_file:
        output_file.write(group_render)

def generate_llms_txt():
    lines = [
        "# Yonatan Bisk / CLAW @ CMU",
        "",
        "> Connecting Language to Action and the World -- Yonatan Bisk's lab at "
        "Carnegie Mellon University, working on embodied AI, grounded language, "
        "robotics, and multimodal machine learning.",
        "",
        "## Pages",
        "",
        "- [Homepage](https://talkingtorobots.com/): About the CLAW lab and its research themes.",
        "- [Publications](https://talkingtorobots.com/publications.html): Full, searchable list of papers.",
        "- [Group](https://talkingtorobots.com/CLAW/): Current students, postdocs, and alumni.",
        "- [Yonatan Bisk](https://talkingtorobots.com/yonatanbisk.html): PI homepage, bio, and contact FAQ.",
        "- [Teaching](https://talkingtorobots.com/teaching.html): Courses taught at CMU.",
        "- [Resources](https://talkingtorobots.com/resources.html): Benchmarks and code released by the lab.",
        "- [Grad School FAQ](https://talkingtorobots.com/FAQ.html): Advice for prospective students.",
        "",
        "## Recent Publications",
        "",
    ]
    for entry in pubs[:10]:
        link = entry.get("url") or "https://talkingtorobots.com/publications.html"
        lines.append(f"- [{entry['title']}]({link}): {entry['venue']} ({entry['year']})")
    lines.append("")

    with open("../llms.txt", 'wt') as output_file:
        output_file.write("\n".join(lines))

def generate_COA():

    ## Create a list of collaborators from the last 48 months with affiliations (for NSF)
    collaborators = {}
    cutoff_year = date.today().year - 4
    for pub in pubs:
        if int(pub["year"]) > cutoff_year and "Open X-" not in pub["title"]:
            for a in pub["authors"]:
              if a != "Yonatan Bisk" and a not in student_names:
                parts = a.rsplit(' ',1)
                last_first = parts[1] + ", "+ parts[0]
                collaborators[last_first] = max(int(pub["year"]), collaborators.get(last_first, 0))

    aff_yaml = load_validated("yaml/affiliations.yaml", Affiliation)
    affiliations = {}
    for entry in aff_yaml:
        affiliations[entry["name"]] = entry["affiliation"]
    
    new_entries = len(collaborators)

    # Load the COA excel spreadsheet
    # Heavily ChatGPT based code here
    workbook = openpyxl.load_workbook('templates/COA.xlsx')
    sheet = workbook.active
    start_row = 52

    # Create a black border style
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Adjust merged cells ranges
    for merge in list(sheet.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(merge))
        if min_row >= start_row:
            sheet.merged_cells.remove(merge)
            new_merge = f"{openpyxl.utils.get_column_letter(min_col)}{min_row + new_entries}:{openpyxl.utils.get_column_letter(max_col)}{max_row + new_entries}"
            sheet.merged_cells.add(new_merge)
    
    # Adjust the TableD5 position (this is the Editorial board)
    table_d5 = None
    for tbl in sheet.tables.values():
        if tbl.name == 'TableD5':
            table_d5 = tbl
            break

    if table_d5:
        min_col, min_row, max_col, max_row = range_boundaries(table_d5.ref)
        new_ref = f"{get_column_letter(min_col)}{min_row + new_entries}:{get_column_letter(max_col)}{max_row + new_entries}"
        table_d5.ref = new_ref

    sheet.insert_rows(start_row, amount=new_entries)

    for i, author in enumerate(collaborators, start=start_row):
        data = ['A:', 
                author, 
                affiliations[author] if author in affiliations else '',
                '',
                '1/1/' + str(collaborators[author])
                ]
        for col_num, value in enumerate(data, start=1):
            cell = sheet.cell(row=i, column=col_num, value=value)
            cell.border = border_style
    workbook.save('current_COA.xlsx')

def main():
    if not args.onepager:
        generate_publications_website()
    generate_CV()
    if not args.onepager:
        generate_group_page()
        generate_COA()
        generate_llms_txt()

if __name__ == "__main__":
    main()
